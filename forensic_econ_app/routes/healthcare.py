from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from forensic_econ_app.models.models import db, Evaluee, HealthcareScenario, MedicalItem, CPIRate
from decimal import Decimal
from datetime import datetime
import decimal
import pandas as pd
import os
from tempfile import NamedTemporaryFile
from flask_login import login_required, current_user
from ..utils.life_care_plan import generate_life_care_plan_table
from openpyxl.utils import get_column_letter

healthcare = Blueprint('healthcare', __name__)

def compute_future_medical_costs(scenario):
    """Compute future medical costs based on scenario parameters."""
    print(f"DEBUG: Starting calculations for scenario {scenario.scenario_name}")
    
    # Determine scenario growth rate (inflation rate)
    if scenario.growth_method == "CPI":
        scenario_growth_rate = CPIRate.get_rate("CPI") or Decimal('0.03')
    elif scenario.growth_method == "PCE":
        scenario_growth_rate = CPIRate.get_rate("PCE") or Decimal('0.025')
    elif scenario.growth_method == "Medical_CPI":
        scenario_growth_rate = CPIRate.get_rate("Medical_CPI") or Decimal('0.035')
    else:
        scenario_growth_rate = scenario.growth_rate_custom

    print(f"DEBUG: Growth rate: {scenario_growth_rate}")

    discount_rate = scenario.discount_rate
    partial_offset = scenario.partial_offset
    total_offset = scenario.total_offset
    projection_years = scenario.projection_years
    discounting_enabled = scenario.discount_method != 'none'

    print(f"DEBUG: Initial parameters - Discount rate: {discount_rate}, Projection years: {projection_years}")

    # If total_offset => discount_rate = growth_rate
    if total_offset:
        discount_rate = scenario_growth_rate

    # partial offset => net_discount in discrete time
    net_discount = discount_rate
    if partial_offset:
        net_discount = ((1 + discount_rate)/(1 + scenario_growth_rate)) - 1
        scenario_growth_rate = Decimal('0.0')  # treat growth as 0, discount = net_discount

    # if discount_method == 'net', treat user-supplied discount_rate as net
    if scenario.discount_method == 'net':
        net_discount = discount_rate

    print(f"DEBUG: Final discount rate: {net_discount}")

    # Helper functions for inflation and discounting
    def inflation_factor(year, inflation_rate):
        """Return the inflation multiplier for a given year (base year = 1)."""
        return (1 + inflation_rate) ** (year - 1)

    def discount_factor(year, discount_rate):
        """Return the discount multiplier for a given year (base year = 1)."""
        return 1 / ((1 + discount_rate) ** (year - 1))

    # Create results dictionary with parameters
    results = {
        "parameters_used": {
            "growth_rate_effective": float(scenario_growth_rate),
            "discount_rate_effective": float(net_discount),
            "projection_years": projection_years,
            "partial_offset": partial_offset,
            "total_offset": total_offset,
            "discounting_enabled": discounting_enabled
        },
        "items_breakdown": [],
        "grand_total_present_value": 0.0,
        "grand_total_undiscounted": 0.0,
        "category_tables": {},
        "category_summaries": []
    }

    # Calculate start age
    if scenario.evaluee.date_of_birth:
        start_date = datetime.now()
        current_age = float(
            (start_date.year - scenario.evaluee.date_of_birth.year) +
            ((start_date.month - scenario.evaluee.date_of_birth.month) / 12.0) +
            ((start_date.day - scenario.evaluee.date_of_birth.day) / 365.25)
        )
    else:
        current_age = 60.39  # Default starting age if no DOB

    current_year = datetime.now().year
    detailed_rows = []

    # Process each medical item
    for item in scenario.medical_items:
        print(f"\nDEBUG: Processing item {item.label}")
        
        try:
            annual_cost = float(item.annual_cost) if item.annual_cost is not None else 0.0
            growth_rate = float(item.growth_rate) if item.growth_rate is not None else float(scenario_growth_rate)
            duration_years = int(item.duration_years) if item.duration_years else projection_years
            start_year = int(item.start_year) if item.start_year is not None else 1
            interval_years = int(item.interval_years) if item.interval_years is not None else 1

            # Calculate yearly projections
            for year in range(start_year, min(start_year + duration_years, projection_years + 1)):
                # Skip years that don't match the interval
                if (year - start_year) % interval_years != 0:
                    continue

                if item.is_one_time and year > start_year:
                    continue

                inf_factor = inflation_factor(year, growth_rate)
                infl_annual_cost = annual_cost * inf_factor
                disc_factor = discount_factor(year, float(net_discount))
                present_val = infl_annual_cost * disc_factor if discounting_enabled else infl_annual_cost
                proj_age = current_age + (year - 1)
                proj_year = current_year + (year - 1)

                detailed_rows.append({
                    "Category": item.category or "Uncategorized",
                    "Service": item.label,
                    "Projection_Year": proj_year,
                    "Age": round(proj_age, 1),
                    "Inflation_Adjusted_Annual_Cost": infl_annual_cost,
                    "Present_Value": present_val
                })

        except (ValueError, TypeError) as e:
            print(f"DEBUG: Error processing item {item.label}: {str(e)}")
            continue

    # Convert to DataFrame for analysis
    detailed_df = pd.DataFrame(detailed_rows)

    # Create pivot tables for each category
    if not detailed_df.empty:
        for cat in detailed_df["Category"].unique():
            cat_df = detailed_df[detailed_df["Category"] == cat]
            
            # Pivot for future values
            pivot_future = cat_df.pivot_table(
                index="Projection_Year", 
                columns="Service",
                values="Inflation_Adjusted_Annual_Cost"
            )
            
            # Insert Age column
            age_series = cat_df.groupby("Projection_Year")["Age"].first()
            pivot_future.insert(0, "Age", age_series)
            
            # Add Total Future Value column
            pivot_future["Total Future Value"] = pivot_future.drop(columns="Age").sum(axis=1)
            
            # Add Total Present Value column
            present_totals = cat_df.groupby("Projection_Year")["Present_Value"].sum()
            pivot_future["Total Present Value"] = present_totals
            
            # Convert index to string and sort
            pivot_future.index = pivot_future.index.map(str)
            pivot_future.sort_index(inplace=True)
            
            # Compute and append totals row
            totals = pivot_future.drop(columns="Age").sum()
            totals["Age"] = "—"
            pivot_future.loc["Total"] = totals
            
            results["category_tables"][cat] = pivot_future.to_dict()

        # Create overall summary by category
        overall_summary = detailed_df.groupby("Category").agg({
            "Inflation_Adjusted_Annual_Cost": "sum",
            "Present_Value": "sum"
        }).reset_index()
        
        overall_summary.rename(columns={
            "Inflation_Adjusted_Annual_Cost": "Total Future Value",
            "Present_Value": "Total Present Value"
        }, inplace=True)
        
        # Add grand totals
        results["grand_total_undiscounted"] = float(overall_summary["Total Future Value"].sum())
        results["grand_total_present_value"] = float(overall_summary["Total Present Value"].sum())
        
        # Store category summaries
        results["category_summaries"] = overall_summary.to_dict(orient="records")

    results["notes"] = (
        "Computed using Eric Christensen's 2022 methodology for life care plan calculations. "
        "Partial/Total offset done in discrete-time form. All currency nominal. "
        "Individual growth rates and age ranges applied where specified."
    )
    
    return results

@healthcare.route('/healthcare/<int:evaluee_id>')
def healthcare_form(evaluee_id):
    """Display healthcare form for an evaluee."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenarios = HealthcareScenario.query.filter_by(evaluee_id=evaluee_id).all()
    cpi_rates = {
        'CPI': CPIRate.get_rate("CPI") or 3.0,
        'PCE': CPIRate.get_rate("PCE") or 2.5,
        'Medical_CPI': CPIRate.get_rate("Medical_CPI") or 3.5
    }
    return render_template('healthcare/form.html', 
                         evaluee=evaluee, 
                         scenarios=scenarios,
                         cpi_rates=cpi_rates,
                         now=datetime.now())

@healthcare.route('/healthcare/<int:evaluee_id>', methods=['POST'])
def create_scenario(evaluee_id):
    """Create a new healthcare scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    
    # Convert percentage inputs to decimal
    growth_rate_custom = Decimal(request.form.get('growth_rate_custom', '3.0')) / 100
    discount_rate = Decimal(request.form.get('discount_rate', '5.0')) / 100
    
    scenario = HealthcareScenario(
        evaluee_id=evaluee_id,
        scenario_name=request.form.get('scenario_name'),
        growth_method=request.form.get('growth_method', 'CPI'),
        growth_rate_custom=growth_rate_custom,
        discount_method=request.form.get('discount_method', 'nominal'),
        discount_rate=discount_rate,
        partial_offset=bool(request.form.get('partial_offset')),
        total_offset=bool(request.form.get('total_offset')),
        projection_years=int(request.form.get('projection_years', 20))
    )
    
    db.session.add(scenario)
    db.session.commit()
    
    flash('Healthcare scenario created successfully.', 'success')
    return redirect(url_for('healthcare.view_scenario', evaluee_id=evaluee_id, scenario_id=scenario.id))

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>')
def view_scenario(evaluee_id, scenario_id):
    """View a healthcare scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HealthcareScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Invalid scenario for this evaluee.', 'error')
        return redirect(url_for('healthcare.healthcare_form', evaluee_id=evaluee_id))
    
    results = compute_future_medical_costs(scenario)
    return render_template('healthcare/view_scenario.html', 
                         evaluee=evaluee, 
                         scenario=scenario, 
                         results=results,
                         now=datetime.now())

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>/edit', methods=['GET', 'POST'])
def edit_scenario(evaluee_id, scenario_id):
    """Edit a healthcare scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HealthcareScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Invalid scenario for this evaluee.', 'error')
        return redirect(url_for('healthcare.healthcare_form', evaluee_id=evaluee_id))
    
    if request.method == 'POST':
        scenario.scenario_name = request.form.get('scenario_name')
        scenario.growth_method = request.form.get('growth_method')
        scenario.growth_rate_custom = Decimal(request.form.get('growth_rate_custom', '3.0')) / 100
        scenario.discount_method = request.form.get('discount_method')
        scenario.discount_rate = Decimal(request.form.get('discount_rate', '5.0')) / 100
        scenario.partial_offset = bool(request.form.get('partial_offset'))
        scenario.total_offset = bool(request.form.get('total_offset'))
        
        # Use life expectancy as default if no projection years provided
        projection_years = request.form.get('projection_years')
        if not projection_years and evaluee.life_expectancy:
            scenario.projection_years = int(evaluee.life_expectancy)
        else:
            scenario.projection_years = int(projection_years or 20)
        
        db.session.commit()
        flash('Healthcare scenario updated successfully.', 'success')
        return redirect(url_for('healthcare.view_scenario', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    cpi_rates = {
        'CPI': CPIRate.get_rate("CPI") or 3.0,
        'PCE': CPIRate.get_rate("PCE") or 2.5,
        'Medical_CPI': CPIRate.get_rate("Medical_CPI") or 3.5
    }
    
    return render_template('healthcare/edit_scenario.html', 
                         evaluee=evaluee, 
                         scenario=scenario,
                         cpi_rates=cpi_rates,
                         now=datetime.now())

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>/items', methods=['GET'])
def manage_items(evaluee_id, scenario_id):
    """Manage medical items for a scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HealthcareScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Invalid scenario for this evaluee.', 'error')
        return redirect(url_for('healthcare.healthcare_form', evaluee_id=evaluee_id))
    
    return render_template('healthcare/items.html', evaluee=evaluee, scenario=scenario)

@healthcare.route('/evaluee/<int:evaluee_id>/healthcare/<int:scenario_id>/items/add', methods=['POST'])
@login_required
def add_item(evaluee_id, scenario_id):
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HealthcareScenario.query.get_or_404(scenario_id)
    
    # Verify ownership
    if evaluee.user_id != current_user.id:
        abort(403)
    if scenario.evaluee_id != evaluee_id:
        abort(400)
    
    # Get form data
    label = request.form.get('label')
    category = request.form.get('category', 'Uncategorized')
    annual_cost = request.form.get('annual_cost', type=float)
    is_one_time = request.form.get('is_one_time') == 'on'
    interval_years = request.form.get('interval_years', type=int, default=1)
    
    # Get growth rate (convert from percentage to decimal)
    growth_rate_str = request.form.get('growth_rate')
    growth_rate = float(growth_rate_str) / 100 if growth_rate_str else None
    
    # Get age ranges (now as floats)
    age_initiated = request.form.get('age_initiated', type=float)
    age_through = request.form.get('age_through', type=float)
    
    # Get year-based timing (now optional)
    start_year = request.form.get('start_year', type=int)
    duration_years = request.form.get('duration_years', type=int)
    
    # Validate inputs
    if not label or annual_cost is None or annual_cost < 0:
        flash('Please provide a valid label and cost.', 'danger')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    if growth_rate is not None and (growth_rate < 0 or growth_rate > 1):
        flash('Growth rate must be between 0% and 100%.', 'danger')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    if age_initiated is not None and age_through is not None and age_initiated > age_through:
        flash('Age initiated must be less than or equal to age through.', 'danger')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    if start_year is not None and start_year < 1:
        flash('Start year must be 1 or greater.', 'danger')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    if duration_years is not None and duration_years < 1:
        flash('Duration must be 1 year or greater.', 'danger')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    if interval_years is not None and interval_years < 1:
        flash('Interval must be 1 year or greater.', 'danger')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    # Create new medical item
    item = MedicalItem(
        label=label,
        category=category,
        annual_cost=annual_cost,
        is_one_time=is_one_time,
        growth_rate=growth_rate,
        age_initiated=age_initiated,
        age_through=age_through,
        start_year=start_year or 1,  # Default to 1 if not specified
        duration_years=duration_years,
        interval_years=interval_years,
        scenario_id=scenario_id
    )
    
    db.session.add(item)
    db.session.commit()
    
    flash(f'Added medical cost: {label}', 'success')
    return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>/items/<int:item_id>/delete', methods=['POST'])
def delete_item(evaluee_id, scenario_id, item_id):
    """Delete a medical item."""
    item = MedicalItem.query.get_or_404(item_id)
    if item.scenario_id != scenario_id:
        flash('Invalid item for this scenario.', 'error')
        return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    db.session.delete(item)
    db.session.commit()
    flash('Medical item deleted successfully.', 'success')
    return redirect(url_for('healthcare.manage_items', evaluee_id=evaluee_id, scenario_id=scenario_id))

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>/duplicate', methods=['POST'])
def duplicate_scenario(evaluee_id, scenario_id):
    """Duplicate a healthcare scenario."""
    original = HealthcareScenario.query.get_or_404(scenario_id)
    if original.evaluee_id != evaluee_id:
        flash('Invalid scenario for this evaluee.', 'error')
        return redirect(url_for('healthcare.healthcare_form', evaluee_id=evaluee_id))
    
    new_scenario = HealthcareScenario(
        evaluee_id=evaluee_id,
        scenario_name=f"{original.scenario_name} (Copy)",
        growth_method=original.growth_method,
        growth_rate_custom=original.growth_rate_custom,
        discount_method=original.discount_method,
        discount_rate=original.discount_rate,
        partial_offset=original.partial_offset,
        total_offset=original.total_offset,
        projection_years=original.projection_years
    )
    db.session.add(new_scenario)
    db.session.flush()  # Get the new scenario ID
    
    # Copy medical items
    for item in original.medical_items:
        new_item = MedicalItem(
            scenario_id=new_scenario.id,
            label=item.label,
            category=item.category,
            annual_cost=item.annual_cost,
            is_one_time=item.is_one_time,
            growth_rate=item.growth_rate,
            age_initiated=item.age_initiated,
            age_through=item.age_through,
            start_year=item.start_year,
            duration_years=item.duration_years,
            interval_years=item.interval_years
        )
        db.session.add(new_item)
    
    db.session.commit()
    flash('Healthcare scenario duplicated successfully.', 'success')
    return redirect(url_for('healthcare.view_scenario', evaluee_id=evaluee_id, scenario_id=new_scenario.id))

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>/delete', methods=['POST'])
def delete_scenario(evaluee_id, scenario_id):
    """Delete a healthcare scenario."""
    scenario = HealthcareScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Invalid scenario for this evaluee.', 'error')
        return redirect(url_for('healthcare.healthcare_form', evaluee_id=evaluee_id))
    
    try:
        db.session.delete(scenario)
        db.session.commit()
        flash('Healthcare scenario deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting healthcare scenario.', 'error')
    
    return redirect(url_for('evaluee.view', evaluee_id=evaluee_id))

@healthcare.route('/healthcare/<int:evaluee_id>/scenario/<int:scenario_id>/export')
def export_scenario(evaluee_id, scenario_id):
    """Export healthcare scenario to Excel."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HealthcareScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Invalid scenario for this evaluee.', 'error')
        return redirect(url_for('healthcare.healthcare_form', evaluee_id=evaluee_id))
    
    results = compute_future_medical_costs(scenario)
    current_year = datetime.now().year
    
    # Create a temporary file
    with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        writer = pd.ExcelWriter(tmp.name, engine='openpyxl')
        
        # Parameters sheet
        params_data = {
            'Parameter': [
                'Scenario Name',
                'Growth Method',
                'Custom Growth Rate' if scenario.growth_method == 'custom' else 'Effective Growth Rate',
                'Discount Method',
                'Discount Rate',
                'Projection Years',
                'Partial Offset',
                'Total Offset'
            ],
            'Value': [
                scenario.scenario_name,
                scenario.growth_method,
                f"{float(scenario.growth_rate_custom * 100 if scenario.growth_method == 'custom' else results['parameters_used']['growth_rate_effective']):.2f}%",
                scenario.discount_method,
                f"{float(scenario.discount_rate * 100):.2f}%",
                scenario.projection_years,
                'Yes' if scenario.partial_offset else 'No',
                'Yes' if scenario.total_offset else 'No'
            ]
        }
        pd.DataFrame(params_data).to_excel(writer, sheet_name='Parameters', index=False)

        # Create category sheets with vertical layout
        for category, table in results['category_tables'].items():
            # Convert the table data to a list format
            data_rows = []
            
            # Add header rows
            data_rows.append(['', '2025 Cost:'])  # First row with initial costs
            data_rows.append(['', 'Duration (Years)'])  # Duration row
            data_rows.append(['', 'Growth Rate:'])  # Growth rate row
            data_rows.append(['', 'Discount Rate:'])  # Discount rate row
            data_rows.append(['', 'Interval (Years):'])  # Interval row
            
            # Get all years (excluding 'Total' row)
            years = [year for year in table['Age'].keys() if year != 'Total']
            
            # Get all services (excluding Age, Total Future Value, and Total Present Value)
            services = [col for col in table.keys() 
                       if col not in ['Age', 'Total Future Value', 'Total Present Value']]
            
            # Add column headers
            headers = ['Year', 'Age'] + services + ['Total Future Value', 'Total Present Value']
            
            # Fill in the initial rows
            for service in services:
                # Add service column to cost row
                data_rows[0].append(f"${table[service]['Total']:,.2f}")
                
                # Add duration
                service_item = next((item for item in scenario.medical_items if item.label == service), None)
                data_rows[1].append(f"{service_item.duration_years or 1:.2f}" if service_item else "1.00")
                
                # Add growth rate
                growth_rate = service_item.growth_rate if service_item and service_item.growth_rate is not None \
                    else results['parameters_used']['growth_rate_effective']
                data_rows[2].append(f"{float(growth_rate * 100):.2f}%")
                
                # Add discount rate
                data_rows[3].append(f"{float(results['parameters_used']['discount_rate_effective'] * 100):.2f}%")
                
                # Add interval
                data_rows[4].append(f"Every {service_item.interval_years} year{'s' if service_item.interval_years != 1 else ''}" if service_item else "Annual")
            
            # Add totals columns to header rows
            for _ in range(2):  # Add empty cells for Total Future Value and Total Present Value
                data_rows[0].append('')
                data_rows[1].append('')
                data_rows[2].append('')
                data_rows[3].append('')
                data_rows[4].append('')
            
            # Add year-by-year data
            for year in sorted(years):
                row = [year]  # Start with the year
                row.append(table['Age'][year])  # Add age
                
                # Add service values
                for service in services:
                    row.append(f"${table[service][year]:,.2f}" if table[service][year] else "$0.00")
                
                # Add totals
                row.append(f"${table['Total Future Value'][year]:,.2f}")
                row.append(f"${table['Total Present Value'][year]:,.2f}")
                data_rows.append(row)
            
            # Add total row
            total_row = ['', 'Total:']
            for service in services:
                total_row.append(f"${table[service]['Total']:,.2f}")
            total_row.append(f"${table['Total Future Value']['Total']:,.2f}")
            total_row.append(f"${table['Total Present Value']['Total']:,.2f}")
            data_rows.append(total_row)
            
            # Create DataFrame and write to Excel
            df = pd.DataFrame(data_rows[1:], columns=headers)  # Skip the first row as it's our header
            df.to_excel(writer, sheet_name=category, index=False)
            
            # Format the worksheet
            worksheet = writer.sheets[category]
            
            # Set column widths
            for col_idx, col_name in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = 20
            
            # Add borders and formatting
            from openpyxl.styles import Border, Side, Alignment, Font
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders and center alignment to all cells
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
            
            # Make headers and first column bold
            for row in worksheet[1:6]:  # First 5 rows (parameters including interval)
                for cell in row:
                    cell.font = Font(bold=True)
            
            # Make the total row bold
            for cell in worksheet[worksheet.max_row]:
                cell.font = Font(bold=True)
        
        # Overall Summary sheet
        summary_df = pd.DataFrame(results['category_summaries'])
        summary_df.to_excel(writer, sheet_name='Overall Summary', index=False)
        
        # Format the summary worksheet
        worksheet = writer.sheets['Overall Summary']
        
        # Set column widths
        for col_idx, col_name in enumerate(summary_df.columns):
            col_letter = get_column_letter(col_idx + 1)
            worksheet.column_dimensions[col_letter].width = 20
        
        # Add borders and formatting
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # Make headers bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        
        # Add grand total row
        grand_total_row = worksheet.max_row + 2
        worksheet.cell(row=grand_total_row, column=1, value='Grand Total')
        worksheet.cell(row=grand_total_row, column=2, value=results['grand_total_undiscounted'])
        worksheet.cell(row=grand_total_row, column=3, value=results['grand_total_present_value'])
        
        # Format grand total row
        for col in range(1, 4):
            cell = worksheet.cell(row=grand_total_row, column=col)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        writer.close()
        
        return send_file(
            tmp.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'healthcare_scenario_{scenario.scenario_name}_{evaluee.last_name}.xlsx'
        ) 