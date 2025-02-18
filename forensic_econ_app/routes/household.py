from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from ..models.models import db, Evaluee, HouseholdServicesScenario, HouseholdServiceStage
from decimal import Decimal
import pandas as pd
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font

household = Blueprint('household', __name__)

@household.route('/household/<int:evaluee_id>')
@login_required
def household_form(evaluee_id):
    """Display the household services form and list existing scenarios."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenarios = HouseholdServicesScenario.query.filter_by(evaluee_id=evaluee_id).all()
    return render_template('household/form.html', evaluee=evaluee, scenarios=scenarios)

@household.route('/household/<int:evaluee_id>', methods=['POST'])
@login_required
def create_scenario(evaluee_id):
    """Create a new household services scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    
    try:
        scenario = HouseholdServicesScenario(
            evaluee_id=evaluee_id,
            scenario_name=request.form['scenario_name'],
            area_wage_adjustment=Decimal(request.form['area_wage_adjustment']) / 100,
            reduction_percentage=Decimal(request.form['reduction_percentage']) / 100,
            growth_rate=Decimal(request.form['growth_rate']) / 100,
            discount_rate=Decimal(request.form['discount_rate']) / 100
        )
        
        db.session.add(scenario)
        db.session.commit()
        
        flash('Scenario created successfully. Add stages to complete the calculation.')
        return redirect(url_for('household.manage_stages', evaluee_id=evaluee_id, scenario_id=scenario.id))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating scenario: {str(e)}')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>')
@login_required
def view_scenario(evaluee_id, scenario_id):
    """View a specific household services scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    # Calculate present value
    scenario.present_value = scenario.calculate_present_value()
    db.session.commit()
    
    return render_template('household/view_scenario.html', evaluee=evaluee, scenario=scenario, Decimal=Decimal)

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>/stages')
@login_required
def manage_stages(evaluee_id, scenario_id):
    """Manage stages for a household services scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    return render_template('household/stages.html', evaluee=evaluee, scenario=scenario)

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>/stages/add', methods=['POST'])
@login_required
def add_stage(evaluee_id, scenario_id):
    """Add a new stage to a scenario."""
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    try:
        stage = HouseholdServiceStage(
            scenario_id=scenario_id,
            stage_number=int(request.form['stage_number']),
            years=int(request.form['years']),
            annual_value=Decimal(request.form['annual_value'])
        )
        
        db.session.add(stage)
        db.session.commit()
        
        # Recalculate present value
        scenario.present_value = scenario.calculate_present_value()
        db.session.commit()
        
        flash('Stage added successfully.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding stage: {str(e)}')
    
    return redirect(url_for('household.manage_stages', evaluee_id=evaluee_id, scenario_id=scenario_id))

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>/stages/<int:stage_id>/delete', methods=['POST'])
@login_required
def delete_stage(evaluee_id, scenario_id, stage_id):
    """Delete a stage from a scenario."""
    stage = HouseholdServiceStage.query.get_or_404(stage_id)
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    try:
        db.session.delete(stage)
        
        # Recalculate present value
        scenario.present_value = scenario.calculate_present_value()
        db.session.commit()
        
        flash('Stage deleted successfully.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting stage: {str(e)}')
    
    return redirect(url_for('household.manage_stages', evaluee_id=evaluee_id, scenario_id=scenario_id))

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_scenario(evaluee_id, scenario_id):
    """Edit a household services scenario."""
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    if request.method == 'POST':
        try:
            scenario.scenario_name = request.form['scenario_name']
            scenario.area_wage_adjustment = Decimal(request.form['area_wage_adjustment']) / 100
            scenario.reduction_percentage = Decimal(request.form['reduction_percentage']) / 100
            scenario.growth_rate = Decimal(request.form['growth_rate']) / 100
            scenario.discount_rate = Decimal(request.form['discount_rate']) / 100
            
            # Recalculate present value
            scenario.present_value = scenario.calculate_present_value()
            db.session.commit()
            
            flash('Scenario updated successfully.')
            return redirect(url_for('household.view_scenario', evaluee_id=evaluee_id, scenario_id=scenario_id))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating scenario: {str(e)}')
    
    return render_template('household/edit_scenario.html', evaluee=evaluee, scenario=scenario)

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>/delete', methods=['POST'])
@login_required
def delete_scenario(evaluee_id, scenario_id):
    """Delete a household services scenario."""
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    try:
        db.session.delete(scenario)
        db.session.commit()
        flash('Scenario deleted successfully.')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting scenario: {str(e)}')
    
    return redirect(url_for('household.household_form', evaluee_id=evaluee_id))

@household.route('/household/<int:evaluee_id>/scenario/<int:scenario_id>/export')
@login_required
def export_scenario(evaluee_id, scenario_id):
    """Export scenario to Excel with corrected annual breakdown and PV calculations."""
    scenario = HouseholdServicesScenario.query.get_or_404(scenario_id)
    evaluee = Evaluee.query.get_or_404(evaluee_id)
    
    if scenario.evaluee_id != evaluee_id:
        flash('Access denied.')
        return redirect(url_for('household.household_form', evaluee_id=evaluee_id))
    
    try:
        # Create Excel file
        temp_dir = os.path.dirname(os.path.abspath(__file__))
        temp_path = os.path.join(temp_dir, f'household_services_{scenario_id}.xlsx')
        
        # Create DataFrame for stages
        stages_df = pd.DataFrame([{
            'Stage': stage.stage_number,
            'Years': stage.years,
            'Annual Value': float(stage.annual_value)
        } for stage in sorted(scenario.stages, key=lambda x: x.stage_number)])
        
        # -------------------------------
        # Corrected Annual Breakdown
        # -------------------------------
        breakdown_data = []
        total_pv = 0.0
        
        # For each stage, reset "local_year" so each stage starts at year 1
        for stage in sorted(scenario.stages, key=lambda x: x.stage_number):
            for local_year in range(1, stage.years + 1):
                # Growth from the stage's base annual_value
                grown_value = float(stage.annual_value) * (
                    (1 + float(scenario.growth_rate)) ** (local_year - 1)
                )
                
                # Apply area wage adjustment & reduction
                adjusted_value = grown_value * float(scenario.area_wage_adjustment) * float(scenario.reduction_percentage)
                
                # Discount from year 1..n (local_year)
                present_value = adjusted_value / (
                    (1 + float(scenario.discount_rate)) ** local_year
                )
                
                total_pv += present_value
                
                breakdown_data.append({
                    'Stage': stage.stage_number,
                    'Stage Year': local_year,
                    'Base Annual Value': float(stage.annual_value),
                    'Grown Value': grown_value,
                    'Adjusted Value': adjusted_value,
                    'Present Value': present_value
                })
        
        # The final present value is simply the sum of discounted, adjusted cash flows
        final_pv = total_pv
        
        breakdown_df = pd.DataFrame(breakdown_data)
        
        # Create Excel writer
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            # Write summary
            summary_df = pd.DataFrame([{
                'Scenario Name': scenario.scenario_name,
                'Area Wage Adjustment': f"{float(scenario.area_wage_adjustment) * 100:.1f}%",
                'Reduction Percentage': f"{float(scenario.reduction_percentage) * 100:.1f}%",
                'Growth Rate': f"{float(scenario.growth_rate) * 100:.1f}%",
                'Discount Rate': f"{float(scenario.discount_rate) * 100:.1f}%",
                'Present Value': final_pv
            }])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the Summary sheet
            worksheet = writer.sheets['Summary']
            for cell in worksheet['F']:  # Present Value column
                if cell.row > 1:  # Skip header
                    cell.number_format = '"$"#,##0.00'
            
            # Write stages with currency formatting
            stages_df.to_excel(writer, sheet_name='Stages', index=False)
            worksheet = writer.sheets['Stages']
            for cell in worksheet['C']:  # Annual Value column
                if cell.row > 1:  # Skip header
                    cell.number_format = '"$"#,##0.00'
            
            # Write annual breakdown with currency formatting
            breakdown_df.to_excel(writer, sheet_name='Annual Breakdown', index=False)
            worksheet = writer.sheets['Annual Breakdown']
            
            # Format currency columns in Annual Breakdown
            currency_columns = ['Base Annual Value', 'Grown Value', 'Adjusted Value', 'Present Value']
            for col_name in currency_columns:
                col_letter = get_column_letter(breakdown_df.columns.get_loc(col_name) + 1)
                for cell in worksheet[col_letter]:
                    if cell.row > 1:  # Skip header
                        cell.number_format = '"$"#,##0.00'
            
            # Add borders and center alignment to all sheets
            for sheet_name in ['Summary', 'Stages', 'Annual Breakdown']:
                worksheet = writer.sheets[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center')
                
                # Make headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
        
        return send_file(
            temp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'household_services_{evaluee.last_name}_{evaluee.first_name}.xlsx'
        )
    
    except Exception as e:
        flash(f'Error exporting scenario: {str(e)}')
        return redirect(url_for('household.view_scenario', evaluee_id=evaluee_id, scenario_id=scenario_id))
    
    finally:
        # Clean up temporary file
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass 